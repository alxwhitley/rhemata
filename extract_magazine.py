#!/usr/bin/env python3
"""
New Wine Magazine Extraction Script
Extracts articles from scanned PDF issues using GPT-4o Vision.
Processes in batches, checkpoints progress, writes structured .txt output.
"""

import os
import re
import sys
import json
import base64
import math
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Tuple

import openai
import fitz  # PyMuPDF
from openpyxl import Workbook, load_workbook
from io import BytesIO
from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parent / "backend" / "app" / ".env")

logger = logging.getLogger(__name__)

# ── CONFIGURATION ────────────────────────────────────────────────────────────

MAGAZINE_DIR = Path("/Users/alexwhitley/Desktop/rhemata/pdf/magazine")
OUTPUT_DIR = Path("/Users/alexwhitley/Desktop/rhemata/pipeline/raw_extracted")
DONE_DIR = Path("/Users/alexwhitley/Desktop/rhemata/pdf/magazine_done")
CHECKPOINT_DIR = Path("/Users/alexwhitley/Desktop/rhemata/pipeline/checkpoints")
TRACKER_PATH = Path("/Users/alexwhitley/Desktop/rhemata/pipeline/rhemata_tracker.xlsx")

BLOCKED_LOG_PATH = Path("/Users/alexwhitley/Desktop/rhemata/pipeline/blocked_batches.json")

MODEL = "gpt-4o"
BATCH_SIZE = 6  # pages per batch
MAX_TOKENS = 8192

# ── CLIENT (lazy init — not needed for dry run) ─────────────────────────────

_client = None

def get_client() -> openai.OpenAI:
    global _client
    if _client is None:
        _client = openai.OpenAI(api_key=os.environ["OPENAI_API_KEY"])
    return _client

# ── SYSTEM PROMPT ────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are processing historical theological magazine content for archival \
research purposes. Extract all text faithfully without omission.

You are extracting articles from a scanned New Wine Magazine PDF — a \
charismatic Christian publication from the 1970s-80s. The pages are \
3-column layout. Articles span columns and continue across pages.

TABLE OF CONTENTS (batch 1 only):
If this is batch 1 and a table of contents page is visible, extract it first:
TABLE OF CONTENTS
- [Article Title] | [Author] | Page [X]
(note if any entry is: editorial, advertisement, bible study, Q&A, \
testimonial, forum)
END TABLE OF CONTENTS

INCLUDE ONLY teaching-focused content. An article qualifies if \
its PRIMARY purpose is to teach a biblical principle, theological \
concept, or spiritual practice. It may use illustrations or \
examples but the core must be instructional.

ALWAYS SKIP — do not extract, do not mention in output:
- Bible study worksheets (fill-in-the-blank, question lists, \
scripture lookup exercises)
- Testimonials and personal salvation stories
- Local church news, event reports, conference recaps
- Historical accounts of what God did in a specific city/region \
even if they contain spiritual language
- Firsthand narratives primarily structured as "here is what \
happened to us/our church"
- Letters to the editor
- Advertisements
- Editorials that are announcements or housekeeping
- Q&A sections and reader questions
- Forum discussions that are primarily conversational exchange \
rather than sustained theological teaching

BORDERLINE CASES — apply this test:
If you removed all the story/narrative/local context, would \
substantial transferable teaching content remain? \
If YES → include. If NO → skip.

Forum and panel discussions: include ONLY if the participants \
are making sustained theological arguments, not just sharing \
opinions or telling stories. If it reads like a conversation, \
skip it. If it reads like teaching in dialogue form, include it.

When skipping content, note it internally as: \
[SKIPPED: reason] \
but do not include these notes in the output file.

FOR EACH qualifying ARTICLE extract the following.

ARTICLE START
TITLE: [title]
AUTHOR: [full name — infer if only surname given, note assumption]
MAGAZINE: New Wine
ISSUE: [issue identifier from filename]
DATE: [Month Year]
CATEGORIES: [comma-separated from: apostolic, church, Christian living, \
community, discipleship, deliverance, family, finances, healing, kingdom, \
leadership, prayer, prophecy, revival, spiritual warfare, worship]
CATEGORY REASONING: [1-2 sentences]

CONTENT:
[Full article text. Preserve subtitles. Fix obvious OCR errors. \
Maintain paragraph breaks. Do not summarize — full text only.]

BOOKS MENTIONED: [Title by Author, or None]
ARTICLE END

If an article is cut off at the end of your pages, write:
[CONTINUED IN NEXT BATCH]

End with:
BOOKS IN THIS BATCH: [Title by Author | one per line, or None]"""


# ── PDF TO IMAGES ────────────────────────────────────────────────────────────

def pdf_to_base64_images(pdf_path: Path) -> List[str]:
    """Convert PDF pages to base64-encoded JPEG images at 300 DPI using PyMuPDF."""
    print(f"  Converting PDF to images at 300 DPI...")
    doc = fitz.open(str(pdf_path))
    # 300 DPI = 300/72 = 4.1667x zoom
    zoom = 300 / 72
    matrix = fitz.Matrix(zoom, zoom)
    images = []
    for page in doc:
        pix = page.get_pixmap(matrix=matrix)
        # Convert to JPEG via PIL for quality control
        img_data = pix.tobytes("jpeg")
        b64 = base64.b64encode(img_data).decode("utf-8")
        images.append(b64)
    doc.close()
    print(f"  {len(images)} page images created")
    return images


# ── METADATA PARSING ─────────────────────────────────────────────────────────

def parse_metadata(filename: str) -> Dict[str, Optional[str]]:
    """Parse issue metadata from filename like NewWineMagazine_Issue_02-1974.pdf"""
    meta = {"issue": None, "year": None, "month": None}
    match = re.search(r'Issue_(\d+)-(\d{4})', filename)
    if match:
        meta["issue"] = match.group(1)
        meta["year"] = match.group(2)
    # Try to extract month if present
    month_match = re.search(r'(\w+)_Issue', filename)
    if month_match:
        meta["month"] = month_match.group(1)
    return meta


# ── CHECKPOINT MANAGEMENT ────────────────────────────────────────────────────

def load_checkpoint(issue_stem: str) -> Optional[Dict]:
    """Load checkpoint file for an issue if it exists."""
    cp_path = CHECKPOINT_DIR / f"{issue_stem}.json"
    if cp_path.exists():
        with open(cp_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def save_checkpoint(issue_stem: str, checkpoint: Dict) -> None:
    """Write checkpoint file to disk."""
    cp_path = CHECKPOINT_DIR / f"{issue_stem}.json"
    with open(cp_path, "w", encoding="utf-8") as f:
        json.dump(checkpoint, f, indent=2)


# ── TRACKER (ATOMIC EXCEL SAVES) ────────────────────────────────────────────

def init_tracker() -> None:
    """Create tracker workbook with headers if it doesn't exist."""
    if TRACKER_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Magazine Issues"
    ws.append([
        "Filename", "Issue", "Year", "Total Pages", "Total Batches",
        "Articles Extracted", "Status", "Output File"
    ])
    wb.save(str(TRACKER_PATH))
    wb.close()
    print(f"  Created tracker: {TRACKER_PATH}")


def update_tracker(
    filename: str,
    issue: Optional[str],
    year: Optional[str],
    total_pages: int,
    total_batches: int,
    articles_extracted: int,
    status: str,
    output_file: str
) -> None:
    """Load workbook, append row, save and close. Never holds workbook open."""
    wb = load_workbook(str(TRACKER_PATH))
    ws = wb["Magazine Issues"]
    ws.append([
        filename, issue, year, total_pages, total_batches,
        articles_extracted, status, output_file
    ])
    wb.save(str(TRACKER_PATH))
    wb.close()


def is_tracked(filename: str) -> bool:
    """Check if a filename is already in the tracker with status 'complete'."""
    if not TRACKER_PATH.exists():
        return False
    wb = load_workbook(str(TRACKER_PATH), read_only=True)
    ws = wb["Magazine Issues"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == filename and row[6] == "complete":
            wb.close()
            return True
    wb.close()
    return False


# ── THEMATIC STITCHING ───────────────────────────────────────────────────────

def extract_partial_context(response_text: str) -> str:
    """Extract the last full paragraph from the last CONTENT: block.
    A full paragraph ends with a period before a double newline or end of string.
    Falls back to last 500 characters if no clean boundary found."""
    # Find the last CONTENT: block
    content_blocks = re.findall(
        r'CONTENT:\s*\n(.*?)(?=\n(?:BOOKS MENTIONED|ARTICLE END|\[CONTINUED))',
        response_text,
        re.DOTALL
    )
    if not content_blocks:
        # Fallback: last 500 chars of entire response
        return response_text.strip()[-500:]

    last_content = content_blocks[-1].strip()

    # Split into paragraphs (double newline separated)
    paragraphs = re.split(r'\n\n+', last_content)

    # Find last paragraph that ends with a period
    for para in reversed(paragraphs):
        para = para.strip()
        if para and para.endswith('.'):
            return para

    # Fallback: last 500 chars of the content block
    return last_content[-500:]


# ── BLOCKED BATCH LOGGING ───────────────────────────────────────────────────

def _load_blocked_log() -> List[Dict]:
    if BLOCKED_LOG_PATH.exists():
        with open(BLOCKED_LOG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def _save_blocked_log(log: List[Dict]) -> None:
    BLOCKED_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(BLOCKED_LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(log, f, indent=2)


# ── GPT-4o VISION API CALL ─────────────────────────────────────────────────

def call_gpt4o(
    images: List[str],
    batch_num: int,
    total_batches: int,
    issue_name: str,
    partial_context: Optional[str] = None,
    start_page: int = 0,
    end_page: int = 0,
) -> str:
    """Send a batch of page images to GPT-4o for extraction."""
    # Build user message content (OpenAI vision format)
    content = []

    # Add bridge context if resuming from a previous batch
    if partial_context and batch_num > 0:
        bridge = (
            f"[BRIDGE CONTEXT - last paragraph of previous batch for continuity]:\n"
            f"{partial_context}\n\n"
            f"This is batch {batch_num + 1} of {total_batches} for issue: {issue_name}.\n"
            f"If the above bridge context ends mid-article, continue that article "
            f"seamlessly before processing new content on these pages."
        )
        content.append({"type": "text", "text": bridge})

    # Add page images (OpenAI base64 image format)
    for img_b64 in images:
        content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/jpeg;base64,{img_b64}",
                "detail": "high",
            }
        })

    if not partial_context or batch_num == 0:
        content.append({
            "type": "text",
            "text": f"This is batch {batch_num + 1} of {total_batches} for issue: {issue_name}. "
                    f"Extract all articles from these pages following the instructions."
        })

    try:
        response = get_client().chat.completions.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": content},
            ],
        )
    except Exception as e:
        logger.exception("GPT-4o extraction failed for batch %d of %s", batch_num + 1, issue_name)
        # Log the blocked batch
        blocked_log = _load_blocked_log()
        blocked_log.append({
            "issue": issue_name,
            "batch_index": batch_num,
            "page_range": [start_page + 1, end_page],
            "error": str(e)[:200],
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        _save_blocked_log(blocked_log)
        raise

    text = response.choices[0].message.content or ""

    if response.choices[0].finish_reason == "length":
        print(f"  WARNING: Batch {batch_num + 1} hit token limit — article may be truncated")
        text += "\n[TRUNCATED - MAX TOKENS HIT]"

    return text


# ── OUTPUT STITCHING ─────────────────────────────────────────────────────────

def stitch_output(batch_outputs: Dict[str, str]) -> str:
    """Merge all batch outputs into a single document.
    Handles [CONTINUED IN NEXT BATCH] markers by joining article fragments."""
    # Sort by batch index
    sorted_keys = sorted(batch_outputs.keys(), key=lambda x: int(x))
    parts = [batch_outputs[k] for k in sorted_keys]

    merged = "\n\n".join(parts)

    # Clean up continuation markers — join split articles
    merged = re.sub(
        r'\[CONTINUED IN NEXT BATCH\]\s*ARTICLE END\s*\n*.*?ARTICLE START\s*\n'
        r'TITLE: \[continued\].*?\nCONTENT:\s*\n?',
        '\n',
        merged,
        flags=re.DOTALL | re.IGNORECASE
    )

    # Also handle simpler continuation patterns
    merged = merged.replace("[CONTINUED IN NEXT BATCH]", "")

    return merged.strip()


def count_articles(text: str) -> int:
    """Count ARTICLE START markers in the output."""
    return len(re.findall(r'ARTICLE START', text))


# ── PROCESS SINGLE ISSUE ────────────────────────────────────────────────────

def process_issue(pdf_path: Path) -> str:
    """Process a single magazine issue. Returns 'processed', 'skipped', or 'failed'."""
    filename = pdf_path.name
    issue_stem = pdf_path.stem
    output_file = OUTPUT_DIR / f"{issue_stem}.txt"

    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")

    # Check tracker
    if is_tracked(filename):
        print(f"  Already tracked as complete — skipping")
        return "skipped"

    # Check checkpoint
    checkpoint = load_checkpoint(issue_stem)
    if checkpoint and checkpoint.get("status") == "complete":
        print(f"  Checkpoint shows complete — skipping")
        return "skipped"

    # Parse metadata
    meta = parse_metadata(filename)
    print(f"  Issue: {meta['issue']}  Year: {meta['year']}")

    # Convert PDF to images
    images = pdf_to_base64_images(pdf_path)
    total_pages = len(images)
    total_batches = math.ceil(total_pages / BATCH_SIZE)
    print(f"  {total_pages} pages → {total_batches} batches of {BATCH_SIZE}")

    # Initialize or resume checkpoint
    if checkpoint and checkpoint.get("status") == "in_progress":
        completed_batches = set(checkpoint.get("completed_batches", []))
        batch_outputs = checkpoint.get("batch_outputs", {})
        partial_context = checkpoint.get("partial_context")
        print(f"  Resuming from checkpoint — {len(completed_batches)}/{total_batches} batches done")
    else:
        completed_batches = set()
        batch_outputs = {}
        partial_context = None
        checkpoint = {
            "filename": filename,
            "total_batches": total_batches,
            "completed_batches": [],
            "batch_outputs": {},
            "partial_context": None,
            "status": "in_progress"
        }
        save_checkpoint(issue_stem, checkpoint)

    # Process each batch
    issue_name = f"{meta.get('issue', '??')}-{meta.get('year', '????')}"

    for batch_idx in range(total_batches):
        if batch_idx in completed_batches:
            print(f"  Batch {batch_idx + 1}/{total_batches} — already done, skipping")
            # Recover partial_context from this batch for next batch
            if str(batch_idx) in batch_outputs:
                partial_context = extract_partial_context(batch_outputs[str(batch_idx)])
            continue

        start_page = batch_idx * BATCH_SIZE
        end_page = min(start_page + BATCH_SIZE, total_pages)
        batch_images = images[start_page:end_page]

        print(f"  Batch {batch_idx + 1}/{total_batches} (pages {start_page + 1}-{end_page})...")

        try:
            response_text = call_gpt4o(
                images=batch_images,
                batch_num=batch_idx,
                total_batches=total_batches,
                issue_name=issue_name,
                partial_context=partial_context,
                start_page=start_page,
                end_page=end_page,
            )
        except Exception as e:
            print(f"  ERROR in batch {batch_idx + 1}: {e}")
            return "failed"

        # Store batch output
        batch_outputs[str(batch_idx)] = response_text
        completed_batches.add(batch_idx)

        # Extract partial context for next batch
        partial_context = extract_partial_context(response_text)

        # Save checkpoint after every successful batch
        checkpoint["completed_batches"] = sorted(list(completed_batches))
        checkpoint["batch_outputs"] = batch_outputs
        checkpoint["partial_context"] = partial_context
        save_checkpoint(issue_stem, checkpoint)

        print(f"  Batch {batch_idx + 1} complete — checkpoint saved")

    # All batches done — stitch output
    print(f"  Stitching {len(batch_outputs)} batches...")
    full_output = stitch_output(batch_outputs)
    article_count = count_articles(full_output)
    print(f"  {article_count} articles extracted")

    # ── OPERATION ORDER (P0 Fix 3) ──
    # 1. Write .txt output file
    try:
        output_file.write_text(full_output, encoding="utf-8")
        print(f"  Output written: {output_file}")
    except Exception as e:
        print(f"  ERROR writing output: {e}")
        return "failed"

    # 2. Update tracker — if this fails, do NOT move PDF
    try:
        update_tracker(
            filename=filename,
            issue=meta.get("issue"),
            year=meta.get("year"),
            total_pages=total_pages,
            total_batches=total_batches,
            articles_extracted=article_count,
            status="complete",
            output_file=str(output_file)
        )
        print(f"  Tracker updated")
    except Exception as e:
        print(f"  ERROR updating tracker: {e}")
        print(f"  PDF left in input folder for retry")
        return "failed"

    # 3. Move PDF to done
    try:
        dest = DONE_DIR / filename
        pdf_path.rename(dest)
        print(f"  PDF moved to: {dest}")
    except Exception as e:
        print(f"  WARNING: Could not move PDF: {e}")

    # 4. Update checkpoint status to complete
    checkpoint["status"] = "complete"
    save_checkpoint(issue_stem, checkpoint)

    print(f"  Done: {filename}")
    return "processed"


# ── MAIN ─────────────────────────────────────────────────────────────────────

def run():
    """Scan magazine directory and process all PDF issues."""
    # Create directories if needed
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    DONE_DIR.mkdir(parents=True, exist_ok=True)
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)

    # Initialize tracker
    init_tracker()

    # Find PDFs to process
    pdfs = sorted(MAGAZINE_DIR.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {MAGAZINE_DIR}")
        return

    print(f"Found {len(pdfs)} PDF(s) to process")

    processed = skipped = failed = 0
    for pdf_path in pdfs:
        result = process_issue(pdf_path)
        if result == "processed":
            processed += 1
        elif result == "skipped":
            skipped += 1
        else:
            failed += 1

    print(f"\n{'='*60}")
    print(f"Done. {processed} processed, {skipped} skipped, {failed} failed.")


def dry_run():
    """Convert page 1 of the first PDF only. No API calls, no file writes."""
    pdfs = sorted(MAGAZINE_DIR.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {MAGAZINE_DIR}")
        return

    pdf_path = pdfs[0]
    print(f"DRY RUN — targeting: {pdf_path.name}")

    doc = fitz.open(str(pdf_path))
    zoom = 300 / 72
    matrix = fitz.Matrix(zoom, zoom)
    page = doc[0]
    pix = page.get_pixmap(matrix=matrix)
    img_data = pix.tobytes("jpeg")
    size_kb = len(img_data) / 1024
    doc.close()

    print(f"  Page 1 image size: {size_kb:.1f} KB (300 DPI JPEG)")
    print(f"  Would process: {pdf_path.name}")
    print(f"DRY RUN COMPLETE")


if __name__ == "__main__":
    if "--dry-run" in sys.argv:
        dry_run()
    else:
        run()
