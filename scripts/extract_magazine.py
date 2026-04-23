#!/usr/bin/env python3
"""
New Wine Magazine Extraction Pipeline (3-pass)

Pass 1: Vision extraction via Gemini Flash 2.0 (full issue, all pages)
Pass 2: Article segmentation via Groq Llama 3.3 70B
Pass 3: QA inspection via Groq Llama 3.3 70B

Input:  sources/magazine/01_to_extract/*.pdf
Output: sources/magazine/02_extracted/{issue_stem}/*.md
"""

import os
import re
import sys
import json
import time
import argparse
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict

import fitz  # PyMuPDF
from PIL import Image
from google import genai
from google.genai import types
from groq import Groq
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

from bible_refs import normalize_refs, _normalize_ref

try:
    import pytesseract
    _HAS_PYTESSERACT = True
except ImportError:
    _HAS_PYTESSERACT = False

load_dotenv(Path(__file__).resolve().parent.parent / "backend" / "app" / ".env")

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# -- CONFIGURATION -----------------------------------------------------------

ROOT = Path(__file__).resolve().parent.parent
TO_EXTRACT_DIR = ROOT / "sources" / "magazine" / "01_to_extract"
EXTRACTED_DIR = ROOT / "sources" / "magazine" / "02_extracted"
PDF_FAILED_DIR = ROOT / "sources" / "magazine" / "06_failed"
TRACKER_PATH = ROOT / "sources" / "magazine" / "rhemata_tracker.xlsx"

GEMINI_MODEL = "gemini-2.5-flash"
GROQ_MODEL = "llama-3.3-70b-versatile"

SAFETY_SETTINGS = [
    types.SafetySetting(category="HARM_CATEGORY_HATE_SPEECH", threshold="BLOCK_NONE"),
    types.SafetySetting(category="HARM_CATEGORY_DANGEROUS_CONTENT", threshold="BLOCK_NONE"),
    types.SafetySetting(category="HARM_CATEGORY_HARASSMENT", threshold="BLOCK_NONE"),
    types.SafetySetting(category="HARM_CATEGORY_SEXUALLY_EXPLICIT", threshold="BLOCK_NONE"),
]

PASS1_SYSTEM_INSTRUCTION = (
    "You are a neutral, literal archival transcription engine. "
    "Transcribe all text exactly as it appears. You are a tool for historical research; "
    "do not filter theological terminology or metaphysical descriptions."
)

MONTH_MAP = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December",
}

VALID_TAGS = {
    "Baptism in the Spirit", "Speaking in Tongues", "Prophetic Ministry",
    "Word of Knowledge", "Word of Wisdom", "Discerning of Spirits",
    "Miracles and Signs", "The Nine Gifts", "Stirring Up Gifts",
    "Moving in the Spirit", "Fruit of the Spirit", "Fresh Anointing",
    "Filling of the Spirit", "Power for Service",
    "Hearing God's Voice", "Dreams and Visions", "Interpreting Your Dreams",
    "Encounters with God", "Divine Appointments", "Supernatural Peace",
    "Manifestations of God", "Intimacy with Jesus", "Atmosphere of Worship",
    "Spiritual Sight", "Knowing God's Heart", "Personal Revelation",
    "Walking in the Spirit", "Led by the Spirit",
    "Intercessory Prayer", "Authority of the Believer",
    "Tearing Down Strongholds", "Resisting the Enemy", "Victory in Christ",
    "Deliverance from Bondage", "Casting Out Demons", "Spiritual Weapons",
    "Breaking Negative Patterns", "Binding and Loosing", "Armor of God",
    "Warfare in Prayer", "Fasting and Prayer", "Protecting Your Mind",
    "Divine Healing", "Praying for the Sick", "Inner Healing",
    "Emotional Wholeness", "Healing of Memories", "Health and Vitality",
    "Overcoming Fear", "Freedom from Anxiety", "Restoration of Soul",
    "Physical Miracles", "The Will to Heal", "Faith for Healing",
    "God's Comfort", "Wholeness in Christ",
    "Biblical Leadership", "Fivefold Ministry", "Apostolic Oversight",
    "Prophetic Direction", "Pastoral Care", "Delegated Authority",
    "Spiritual Covering", "Accountability in Leadership",
    "Covenant Relationships", "Mentoring Relationships",
    "Leading with Integrity", "Servant Leadership", "Team Ministry",
    "Equipping the Saints", "Elders and Deacons",
    "Spiritual Maturity", "Walking with God", "Discipleship and Mentoring",
    "Accountability in Christ", "Knowing God's Will", "Character of Christ",
    "Honoring Biblical Authority", "Submission to God",
    "Faith and Perseverance", "Stewardship and Finances",
    "Spiritual Disciplines", "Dying to Self", "Holiness and Sanctification",
    "Body Ministry",
    "Kingdom of God", "Word and Spirit", "Biblical Authority",
    "The New Covenant", "The Lordship of Christ", "Grace and Mercy",
    "Salvation and Repentance", "End Times Prophecy", "The Rapture",
    "Second Coming", "The Trinity", "Blood of Jesus", "Heaven and Eternity",
    "Restoration of All Things",
    "Biblical Marriage", "Christian Parenting", "Family Life",
    "Relationship Restoration", "Communication in Marriage",
    "Raising Godly Children", "Singleness and Purity",
    "Friendship in Christ", "Honoring Your Parents", "Forgiving Others",
    "Love and Sacrifice", "Conflict Resolution", "The Christian Home",
}

# -- CLIENTS -----------------------------------------------------------------

_gemini_client = None
_groq_client = None


def get_gemini():
    global _gemini_client
    if _gemini_client is None:
        _gemini_client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])
    return _gemini_client


def reset_gemini():
    """Force a fresh Gemini client on next get_gemini() call."""
    global _gemini_client
    _gemini_client = None


def get_groq():
    global _groq_client
    if _groq_client is None:
        _groq_client = Groq(api_key=os.environ["GROQ_API_KEY"])
    return _groq_client


# -- METADATA PARSING --------------------------------------------------------

def parse_issue_meta(filename: str) -> Dict[str, Optional[str]]:
    """Parse issue/year from filename like NewWineMagazine_Issue_02-1974.pdf"""
    match = re.search(r"Issue_(\d+)-(\d{4})", filename)
    if match:
        month_num = match.group(1)
        year = match.group(2)
        issue = f"{month_num}-{year}"
        month_name = MONTH_MAP.get(month_num, month_num)
        return {"issue": issue, "year": year, "month_num": month_num, "date": f"{month_name} {year}"}
    return {"issue": None, "year": None, "month_num": None, "date": None}


def slugify(text: str) -> str:
    """Convert title to a filename-safe slug."""
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "_", text)
    return text[:60].rstrip("_")


# -- TRACKER -----------------------------------------------------------------

def init_tracker() -> None:
    if not TRACKER_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Magazine Issues"
        ws.append([
            "Filename", "Issue", "Year", "Pages", "Pass1", "Pass2", "Pass3",
            "Articles", "Pass_Count", "Warn_Count", "Flag_Count", "Status",
        ])
        wb.save(str(TRACKER_PATH))
        wb.close()
        return
    EXPECTED_HEADERS = [
        "Filename", "Issue", "Year", "Pages", "Pass1", "Pass2", "Pass3",
        "Articles", "Pass_Count", "Warn_Count", "Flag_Count", "Status",
    ]
    wb = load_workbook(str(TRACKER_PATH))
    if "Extraction" not in wb.sheetnames:
        ws = wb.create_sheet("Extraction")
        ws.append(EXPECTED_HEADERS)
        wb.save(str(TRACKER_PATH))
    wb.close()


_REVIEW_HEADERS = ["Issue", "Article Title", "Failure Reason", "Output Folder", "Timestamp"]


def _ensure_review_sheet() -> None:
    """Create the 'Review Needed' sheet in the tracker if it doesn't exist."""
    wb = load_workbook(str(TRACKER_PATH))
    if "Review Needed" not in wb.sheetnames:
        ws = wb.create_sheet("Review Needed")
        ws.append(_REVIEW_HEADERS)
        wb.save(str(TRACKER_PATH))
    wb.close()


def log_failure(issue: str, article_title: str, reason: str,
                issue_dir: Path) -> None:
    """Log a failure to the 'Review Needed' tracker sheet and review_needed.log."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    output_folder = str(issue_dir)

    # Append to tracker sheet
    try:
        _ensure_review_sheet()
        wb = load_workbook(str(TRACKER_PATH))
        ws = wb["Review Needed"]
        ws.append([issue, article_title, reason, output_folder, timestamp])
        wb.save(str(TRACKER_PATH))
        wb.close()
    except Exception as exc:
        logger.warning("Failed to write to Review Needed sheet: %s", exc)

    # Append to review_needed.log in the issue folder
    try:
        issue_dir.mkdir(parents=True, exist_ok=True)
        log_path = issue_dir / "review_needed.log"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {article_title} | {reason}\n")
    except Exception as exc:
        logger.warning("Failed to write review_needed.log: %s", exc)


def update_tracker_row(filename: str, data: Dict) -> None:
    """Update or append a row in the tracker for the given filename."""
    wb = load_workbook(str(TRACKER_PATH))
    ws = wb["Extraction"]

    # Find existing row
    target_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value == filename:
            target_row = row_idx
            break

    col_map = {}
    for idx, cell in enumerate(ws[1], start=1):
        col_map[cell.value] = idx

    if target_row:
        for key, val in data.items():
            if key in col_map:
                ws.cell(row=target_row, column=col_map[key], value=val)
    else:
        row_data = [None] * len(col_map)
        row_data[col_map["Filename"] - 1] = filename
        for key, val in data.items():
            if key in col_map:
                row_data[col_map[key] - 1] = val
        ws.append(row_data)

    wb.save(str(TRACKER_PATH))
    wb.close()


# -- PASS 1: VISION EXTRACTION (Gemini Flash 2.0) ----------------------------

PASS1_PROMPT = """Transcribe every word exactly as printed on each page of this magazine scan.

Rules:
- Transcribe every word exactly as printed. Do not paraphrase, summarize, or add any text \
that is not on the page.
- For advertisement pages, set is_advertisement to true and leave content empty.
- For cover pages, back covers, full-page illustrations with no text, order forms, \
and subscription cards, set is_advertisement to true and leave content empty.
- Include: article text, section headings, author names, pull quotes, sidebars, \
letters to editor, editorial, table of contents.
- Preserve paragraph breaks with a blank line in the content string.
- When a page contains two or more SEPARATE articles sharing the same page, transcribe \
the page COLUMN BY COLUMN, left to right. Finish the entire left column before starting \
the right column. Insert the marker === COLUMN BREAK === between columns.
- Do not add any commentary or notes.

Return a JSON object with this exact structure:
{"pages": [{"page_number": <int>, "content": "<transcribed text>", "is_advertisement": <bool>}]}

You MUST return one entry in the pages array for every page image provided, even if the \
page is an advertisement (use empty content for those).

EXAMPLES of correct transcription for pages with theological content:

Example 1 input: A magazine page discussing spiritual warfare.
Example 1 output:
{"pages": [{"page_number": 7, "content": "## Spiritual Warfare: The Battle for the Mind\n\nThe apostle Paul reminds us that we wrestle not against flesh and blood, but against principalities, against powers, against the rulers of the darkness of this world, against spiritual wickedness in high places (Ephesians 6:12). As believers, we must learn to wage war in the spiritual realm using the weapons God has provided.", "is_advertisement": false}]}

Example 2 input: A magazine page discussing deliverance ministry.
Example 2 output:
{"pages": [{"page_number": 14, "content": "## Freedom from Demonic Influence\n\nJesus demonstrated authority over demons throughout His earthly ministry. In Mark 5, He confronted the man possessed by a legion of unclean spirits, commanding them to come out. This encounter reveals that demonic bondage is real, but the power of Christ is greater. The ministry of deliverance remains vital for the church today.", "is_advertisement": false}]}"""

PASS1_BATCH_SIZE = 5

PASS1_CONFIG = types.GenerateContentConfig(
    system_instruction=PASS1_SYSTEM_INSTRUCTION,
    safety_settings=SAFETY_SETTINGS,
    response_mime_type="application/json",
)

# Layout continuation markers to strip from Pass 1 output.
# Preserves authorial "To be continued ..." and mid-sentence uses of "continued".
# Pass A removes "(Continued on next page)" plus the orphan folio number on the following line.
_CONT_MARKER_WITH_FOLIO = re.compile(
    r"^[ \t]*\(?[ \t]*[Cc]ontinued[ \t]+on[ \t]+next[ \t]+page[ \t]*\)?[ \t]*\n"
    r"[ \t]*\d+[ \t]*(?=\n|$)",
    re.MULTILINE,
)
# Pass B removes standalone markers: "(Continued on/from page N)", "(continued on pg. N)",
# bare "Continued from page N", and any remaining "(Continued on next page)".
_CONT_MARKER = re.compile(
    r"^[ \t]*\(?[ \t]*[Cc]ontinued[ \t]+"
    r"(?:(?:on|from)[ \t]+(?:page|pg\.?)[ \t]+\d+|on[ \t]+next[ \t]+page)"
    r"[ \t]*\)?[ \t]*(?=\n|$)",
    re.MULTILINE,
)


# Bible reference regex: matches "Book Chapter:Verse" patterns inline.
# Handles numbered books (1/2/3), abbreviated and full names, verse ranges.
_BIBLE_BOOK_NAMES = (
    r"(?:(?:[1-3]\s*)?(?:"
    r"Genesis|Exodus|Leviticus|Numbers|Deuteronomy|Joshua|Judges|Ruth|"
    r"Samuel|Kings|Chronicles|Ezra|Nehemiah|Esther|Job|Psalms?|Proverbs|"
    r"Ecclesiastes|Song of Solomon|Isaiah|Jeremiah|Lamentations|Ezekiel|"
    r"Daniel|Hosea|Joel|Amos|Obadiah|Jonah|Micah|Nahum|Habakkuk|"
    r"Zephaniah|Haggai|Zechariah|Malachi|"
    r"Matthew|Mark|Luke|John|Acts|Romans|Corinthians|Galatians|"
    r"Ephesians|Philippians|Colossians|Thessalonians|Timothy|Titus|"
    r"Philemon|Hebrews|James|Peter|Jude|Revelation|"
    # Common abbreviations
    r"Gen|Exo?|Lev|Num|Deut?|Josh|Judg|Sam|Kgs|Chr|Neh|Est|"
    r"Psa?|Prov?|Eccl?|Isa|Jer|Lam|Ezek?|Dan|Hos|Mic|Hab|Zeph|"
    r"Zech|Mal|Matt?|Mk|Lk|Jn|Rom|Cor|Gal|Eph|Phil|Col|"
    r"Thess?|Tim|Tit|Phm|Heb|Jas|Pet|Rev"
    r")\.?)"
)
_BIBLE_REF_RE = re.compile(
    _BIBLE_BOOK_NAMES + r"\s+(\d+)(?:[:\.](\d+[\-,\s\d]*))?",
    re.IGNORECASE,
)


def _extract_bible_refs_regex(text: str) -> List[str]:
    """Extract Bible references from text using regex only (no LLM).
    Returns normalized, deduped list."""
    raw_refs = []
    for m in _BIBLE_REF_RE.finditer(text):
        raw_refs.append(m.group(0).strip().rstrip("."))
    return normalize_refs(raw_refs)


def _scrub_continuation_markers(text: str) -> str:
    """Strip layout continuation markers from Gemini Pass 1 output."""
    text = _CONT_MARKER_WITH_FOLIO.sub("", text)
    text = _CONT_MARKER.sub("", text)
    return text


def _check_finish_reason(response, label: str) -> bool:
    """Log the finish_reason from Gemini response candidates.
    Returns True if a SAFETY block was detected."""
    safety_blocked = False
    try:
        if response.candidates:
            for i, cand in enumerate(response.candidates):
                if cand.finish_reason and cand.finish_reason != "STOP":
                    logger.warning(
                        "%s — candidate %d finish_reason: %s",
                        label, i, cand.finish_reason,
                    )
                    if cand.finish_reason == "SAFETY":
                        safety_blocked = True
    except Exception:
        pass
    return safety_blocked


def _call_gemini_batch(client, prompt: str, batch_images: list,
                       label: str) -> tuple:
    """Call Gemini with PASS1_CONFIG, parse JSON pages array.
    Returns (list_of_page_dicts_or_None, safety_triggered_bool)."""
    content = [prompt] + list(batch_images)
    response = client.models.generate_content(
        model=GEMINI_MODEL, contents=content, config=PASS1_CONFIG,
    )
    safety_triggered = _check_finish_reason(response, label)

    raw = response.text or ""
    if not raw.strip():
        logger.warning("%s — Gemini returned empty response", label)
        return None, safety_triggered

    try:
        parsed = json.loads(raw)
        pages = parsed.get("pages", [])
        if not pages:
            logger.warning("%s — JSON parsed but 'pages' array is empty", label)
            return None, safety_triggered
        return pages, safety_triggered
    except (json.JSONDecodeError, ValueError) as exc:
        logger.warning("%s — JSON parse failed: %s", label, exc)
        return None, safety_triggered


def _ocr_fallback_page(client, image, page_num: int,
                       pdf_name: str) -> Optional[dict]:
    """Fall back to pytesseract OCR + Gemini text-only cleanup for a single page.
    Returns a page dict or None."""
    if not _HAS_PYTESSERACT:
        logger.warning(
            "Page %d of %s: pytesseract not installed — skipping OCR fallback",
            page_num, pdf_name,
        )
        return None

    try:
        ocr_text = pytesseract.image_to_string(image)
    except Exception as exc:
        logger.warning(
            "Page %d of %s: pytesseract failed: %s", page_num, pdf_name, exc,
        )
        return None

    if len(ocr_text.strip()) < 50:
        logger.warning(
            "Page %d of %s: OCR returned only %d chars — skipping",
            page_num, pdf_name, len(ocr_text.strip()),
        )
        return None

    # Send OCR text to Gemini text-only (no image) for cleanup
    cleanup_prompt = (
        "Fix OCR errors in this archival magazine page transcription. "
        "Preserve all theological terminology exactly as written — do not filter or censor "
        "any religious, spiritual warfare, or metaphysical language. "
        "Return JSON only:\n"
        f'{{"pages": [{{"page_number": {page_num}, "content": "<cleaned text>", '
        f'"is_advertisement": false}}]}}\n\n'
        f"Raw OCR text:\n{ocr_text}"
    )

    label = f"OCR-cleanup p{page_num} of {pdf_name}"
    try:
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[cleanup_prompt],
            config=PASS1_CONFIG,
        )
        safety_hit = _check_finish_reason(response, label)
        if safety_hit:
            # Even text-only got blocked — return raw OCR as best effort
            logger.warning(
                "Page %d of %s: OCR cleanup also hit SAFETY — using raw OCR",
                page_num, pdf_name,
            )
            return {"page_number": page_num, "content": ocr_text.strip(),
                    "is_advertisement": False}

        raw = response.text or ""
        if raw.strip():
            parsed = json.loads(raw)
            pages = parsed.get("pages", [])
            if pages:
                logger.info(
                    "Page %d of %s: vision failed, used pytesseract+Gemini text fallback",
                    page_num, pdf_name,
                )
                return pages[0]
    except Exception as exc:
        logger.warning(
            "Page %d of %s: OCR cleanup failed: %s — using raw OCR",
            page_num, pdf_name, exc,
        )
        return {"page_number": page_num, "content": ocr_text.strip(),
                "is_advertisement": False}

    # Gemini returned empty even on text-only — use raw OCR
    logger.warning(
        "Page %d of %s: OCR cleanup returned empty — using raw OCR",
        page_num, pdf_name,
    )
    return {"page_number": page_num, "content": ocr_text.strip(),
            "is_advertisement": False}


def _extract_single_page(client, image, page_num: int,
                         pdf_name: str) -> Optional[dict]:
    """Retry a single page extraction via vision, falling back to OCR if needed.
    Returns a page dict or None."""
    label = f"single-page retry p{page_num} of {pdf_name}"
    prompt = PASS1_PROMPT + (
        f"\n\nYou are processing page {page_num} only. "
        f"Return exactly one entry with page_number {page_num}."
    )
    pages, safety_triggered = _call_gemini_batch(client, prompt, [image], label)

    if pages:
        return pages[0]

    # Vision failed or got safety-blocked — try OCR fallback
    if safety_triggered:
        logger.warning(
            "Page %d of %s: SAFETY block on vision — attempting OCR fallback",
            page_num, pdf_name,
        )
    return _ocr_fallback_page(client, image, page_num, pdf_name)


def pass1_extract(pdf_path: Path, issue_dir: Path) -> int:
    """Extract raw text from PDF via Gemini Vision in batches. Returns page count."""
    print(f"  PASS 1: Vision extraction via Gemini...")

    doc = fitz.open(str(pdf_path))
    page_count = len(doc)
    images = []
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        images.append(img)
    doc.close()
    print(f"  {page_count} pages converted at 200 DPI")

    client = get_gemini()
    # Collect page dicts keyed by page number
    extracted_pages = {}  # type: Dict[int, dict]

    for batch_start in range(0, page_count, PASS1_BATCH_SIZE):
        batch_end = min(batch_start + PASS1_BATCH_SIZE, page_count)
        batch_images = images[batch_start:batch_end]
        page_num_start = batch_start + 1
        page_num_end = batch_end
        expected_count = batch_end - batch_start

        print(f"  Batch: pages {page_num_start}-{page_num_end}...")

        prompt = PASS1_PROMPT + (
            f"\n\nYou are processing pages {page_num_start} through {page_num_end}. "
            f"Return exactly {expected_count} entries with page_number "
            f"{page_num_start} through {page_num_end}."
        )

        label = f"batch p{page_num_start}-{page_num_end} of {pdf_path.name}"
        pages, safety_triggered = _call_gemini_batch(client, prompt, batch_images, label)

        if pages is None and not safety_triggered:
            # Full batch failed (not safety) — retry once
            logger.warning("%s — retrying full batch", label)
            pages, safety_triggered = _call_gemini_batch(
                client, prompt, batch_images, label + " (retry)",
            )

        # Determine which pages need single-page retry
        needs_single_retry = False
        if safety_triggered:
            logger.warning(
                "SAFETY block on %s — splitting to single-page retries",
                label,
            )
            needs_single_retry = True

        if pages:
            for p in pages:
                pn = p.get("page_number")
                if pn is not None:
                    extracted_pages[pn] = p
            returned = len(pages)
            if returned < expected_count:
                logger.warning(
                    "%s — got %d/%d pages, retrying missing pages individually",
                    label, returned, expected_count,
                )
                needs_single_retry = True
        else:
            needs_single_retry = True

        if needs_single_retry:
            returned_nums = {p.get("page_number") for p in pages} if pages else set()
            for pn in range(page_num_start, page_num_end + 1):
                if pn not in returned_nums and pn not in extracted_pages:
                    logger.info("  Retrying page %d individually...", pn)
                    result = _extract_single_page(
                        client, images[pn - 1], pn, pdf_path.name,
                    )
                    if result:
                        extracted_pages[pn] = result
                    elif safety_triggered:
                        issue_id = parse_issue_meta(pdf_path.name).get("issue", pdf_path.stem)
                        log_failure(issue_id, f"Page {pn}", "SAFETY block after all retries", issue_dir)

        # Reset client after safety trigger to avoid context poisoning
        if safety_triggered:
            logger.info("Safety trigger detected — resetting Gemini client for next batch")
            reset_gemini()
            client = get_gemini()

    # -- Page gap check --
    missing = [pn for pn in range(1, page_count + 1) if pn not in extracted_pages]
    if missing:
        logger.warning(
            "Missing pages after batch extraction: %s — retrying individually",
            missing,
        )
        for pn in missing:
            result = _extract_single_page(client, images[pn - 1], pn, pdf_path.name)
            if result:
                extracted_pages[pn] = result
            else:
                logger.warning("Page %d of %s could not be extracted", pn, pdf_path.name)
                issue_id = parse_issue_meta(pdf_path.name).get("issue", pdf_path.stem)
                log_failure(issue_id, f"Page {pn}", "Empty output after all retries", issue_dir)

    # -- Hard checksum: abort if pages still missing after all retries --
    still_missing = [pn for pn in range(1, page_count + 1) if pn not in extracted_pages]
    if still_missing:
        logger.error(
            "ABORTING %s — %d page(s) still missing after all retries: %s",
            pdf_path.name, len(still_missing), still_missing,
        )
        issue_id = parse_issue_meta(pdf_path.name).get("issue", pdf_path.stem)
        log_failure(
            issue_id, "(whole issue)",
            f"Missing {len(still_missing)} page(s) after all retries: {still_missing}",
            issue_dir,
        )
        # Still save partial raw_text for debugging
        text_parts = []
        for pn in sorted(extracted_pages.keys()):
            page = extracted_pages[pn]
            content = (page.get("content") or "").strip()
            if page.get("is_advertisement", False) or not content:
                continue
            text_parts.append(f"=== PAGE {pn} ===")
            text_parts.append(content)
        issue_dir.mkdir(parents=True, exist_ok=True)
        (issue_dir / "raw_text.txt").write_text("\n\n".join(text_parts), encoding="utf-8")
        return -1

    # -- Assemble raw_text.txt in the legacy === PAGE N === format --
    text_parts = []
    for pn in sorted(extracted_pages.keys()):
        page = extracted_pages[pn]
        content = (page.get("content") or "").strip()
        if page.get("is_advertisement", False) or not content:
            continue
        text_parts.append(f"=== PAGE {pn} ===")
        text_parts.append(content)

    raw_text = "\n\n".join(text_parts)

    # Save raw transcription
    issue_dir.mkdir(parents=True, exist_ok=True)
    (issue_dir / "raw_text.txt").write_text(raw_text, encoding="utf-8")
    print(f"  Raw text saved ({len(raw_text)} chars, {len(extracted_pages)}/{page_count} pages extracted)")

    return page_count


# -- PASS 2: ARTICLE SEGMENTATION (Groq Llama 3.3 70B) ----------------------

PASS2_TOC_SYSTEM = """You are an expert magazine editor. You will be given the full \
transcribed text of a New Wine Magazine issue and its table of contents.

Your job is to identify each article and return ONLY a metadata index. \
Do NOT include article body text.

Rules:
- Use the table of contents as ground truth - find exactly those articles
- Do not include: letters to editor, order forms, subscription info, \
staff boxes, ads, table of contents itself
- Exclude any article that is a Bible Study, Bible lesson, Scripture study, or study guide. \
These are reference materials not theological teaching articles. If an article title contains \
'Bible Study', 'Bible Lesson', or 'Study Guide' skip it entirely.
- Do include: main articles, editorial, forum sections
- Look for "Continued on page X" and "Continued from page X" markers in the text. If an \
article starts on pages 4-5 but has a continuation on page 18, include ALL pages in the \
source_pages list: [4, 5, 18].
- Output as JSON array with this structure:
  [
    {"title": "Article Title", "author": "Author Name", "page_start": 4, "page_end": 10, \
"source_pages": [4, 5, 6, 7, 8, 9, 10]}
  ]
- page_start and page_end refer to the primary === PAGE N === markers in the text
- source_pages is the full list of pages containing content for this article, \
including any continuation pages
- Return ONLY the JSON array, nothing else"""

PASS2_BODY_SYSTEM = """You are an expert magazine transcription editor. You will be \
given a section of raw transcribed magazine text that contains one specific article.

Your job is to extract and clean up ONLY the article specified, formatting it as markdown, \
and assign topic tags from the taxonomy below.

If this article is a Bible Study, Bible lesson, or study guide, \
return an empty JSON object: {} and do not extract it.

Rules:
- Extract the full article text for the specified title and author
- Do NOT include the article title or author name at the start — these are already in metadata
- Start the body text directly with the first paragraph of content
- Do not add, invent, or paraphrase any text
- Do not include text from other articles that may appear in the page range
- If you see a [CONTINUED] marker in the input, it separates non-contiguous page spans of the same article that were stitched together because the article jumped pages. Treat the surrounding text as one continuous article, do not include the [CONTINUED] marker in your output, and bridge across it naturally.
- The raw text may contain === COLUMN BREAK === markers indicating a multi-column page layout. These are layout markers only — do not include them in your output. The article you are extracting may continue after a === COLUMN BREAK === marker; follow the content belonging to this article across column breaks, ignoring any content that belongs to other articles.
- Format as markdown:
  - Section headings as ## H2
  - Any quoted scripture passages that are indented or set apart visually as > blockquote
  - Pull quotes as > *italic blockquote*
  - Normal paragraphs separated by blank lines
  - No H1 - that will be the title

After extracting the article body, assign 5-8 topic tags from the taxonomy below.

STRICT RULES for assigning tags:
- Only assign a tag if the article DIRECTLY teaches on that topic for at least one full paragraph
- Do NOT assign a tag for:
  - Passing mentions or single sentence references
  - Historical or biographical context
  - Tangential connections
  - Topics that are merely implied but not taught
- Ask yourself: 'Would a reader searching for this topic find substantial, helpful content on it \
in this article?' If no, do not assign the tag.
- It is better to assign 3 highly accurate tags than 8 loosely related ones
- Never assign a tag just because a word in the tag appears in the article
- You MUST only return tags from this exact list. Do not create new tags. Do not modify tag names. \
Copy them exactly.

TAXONOMY: Baptism in the Spirit, Speaking in Tongues, Prophetic Ministry, \
Word of Knowledge, Word of Wisdom, Discerning of Spirits, Miracles and Signs, \
The Nine Gifts, Stirring Up Gifts, Moving in the Spirit, Fruit of the Spirit, \
Fresh Anointing, Filling of the Spirit, Power for Service, Hearing God's Voice, \
Dreams and Visions, Interpreting Your Dreams, Encounters with God, Divine Appointments, \
Supernatural Peace, Manifestations of God, Intimacy with Jesus, Atmosphere of Worship, \
Spiritual Sight, Knowing God's Heart, Personal Revelation, Walking in the Spirit, \
Led by the Spirit, Intercessory Prayer, Authority of the Believer, \
Tearing Down Strongholds, Resisting the Enemy, Victory in Christ, \
Deliverance from Bondage, Casting Out Demons, Spiritual Weapons, \
Breaking Negative Patterns, Binding and Loosing, Armor of God, Warfare in Prayer, \
Fasting and Prayer, Protecting Your Mind, Divine Healing, Praying for the Sick, \
Inner Healing, Emotional Wholeness, Healing of Memories, Health and Vitality, \
Overcoming Fear, Freedom from Anxiety, Restoration of Soul, Physical Miracles, \
The Will to Heal, Faith for Healing, God's Comfort, Wholeness in Christ, \
Biblical Leadership, Fivefold Ministry, Apostolic Oversight, Prophetic Direction, \
Pastoral Care, Delegated Authority, Spiritual Covering, Accountability in Leadership, \
Covenant Relationships, Mentoring Relationships, Leading with Integrity, \
Servant Leadership, Team Ministry, Equipping the Saints, Elders and Deacons, \
Spiritual Maturity, Walking with God, Discipleship and Mentoring, \
Accountability in Christ, Knowing God's Will, Character of Christ, \
Honoring Biblical Authority, Submission to God, Faith and Perseverance, \
Stewardship and Finances, Spiritual Disciplines, Dying to Self, \
Holiness and Sanctification, Body Ministry, Kingdom of God, Word and Spirit, \
Biblical Authority, The New Covenant, The Lordship of Christ, Grace and Mercy, \
Salvation and Repentance, End Times Prophecy, The Rapture, Second Coming, \
The Trinity, Blood of Jesus, Heaven and Eternity, Restoration of All Things, \
Biblical Marriage, Christian Parenting, Family Life, Relationship Restoration, \
Communication in Marriage, Raising Godly Children, Singleness and Purity, \
Friendship in Christ, Honoring Your Parents, Forgiving Others, Love and Sacrifice, \
Conflict Resolution, The Christian Home

Return your response as JSON only:
{
  "topic_tags": ["Tag One", "Tag Two", ...],
  "body": "The full formatted article body text..."
}"""


def _extract_toc(raw_text: str) -> str:
    """Extract table of contents section from raw text."""
    # Look for PAGE 3 area where TOC usually lives
    toc_lines = []
    in_toc = False
    for line in raw_text.split("\n"):
        if "=== PAGE 3 ===" in line or "=== PAGE 2 ===" in line:
            in_toc = True
            continue
        if in_toc and line.startswith("=== PAGE"):
            break
        if in_toc:
            toc_lines.append(line)
    return "\n".join(toc_lines).strip() if toc_lines else "(No table of contents found)"


def _extract_page_range(raw_text: str, page_start: int, page_end: int) -> str:
    """Extract text between === PAGE page_start === and === PAGE page_end+1 ===."""
    lines = raw_text.split("\n")
    result = []
    capturing = False
    for line in lines:
        page_match = re.match(r"=== PAGE (\d+) ===", line)
        if page_match:
            page_num = int(page_match.group(1))
            if page_num >= page_start and page_num <= page_end:
                capturing = True
                continue
            elif page_num > page_end:
                break
        if capturing:
            result.append(line)
    return "\n".join(result).strip()


# -- CONTINUATION RESOLUTION ------------------------------------------------

# Case-insensitive: "continued on page N", "continued on pg N", "continued on p. N"
_CONT_ON_REF = re.compile(
    r"continued\s+on\s+(?:page|pg\.?|p\.)\s+(\d+)",
    re.IGNORECASE,
)
# Case-insensitive: "continued from page N", "continued from pg N", "continued from p. N"
_CONT_FROM_REF = re.compile(
    r"continued\s+from\s+(?:page|pg\.?|p\.)\s+(\d+)",
    re.IGNORECASE,
)
_PAGE_HEADER = re.compile(r"=== PAGE (\d+) ===")


def resolve_continuations(raw_text: str) -> Dict[int, List[int]]:
    """Scan raw_text for continuation markers and build a source_page -> [dest_pages] map.

    Uses === PAGE N === headers to determine which page each marker lives on.
    Both "continued on page X" (current -> X) and "continued from page X" (X -> current)
    markers contribute to the map. Destinations are deduplicated in insertion order.
    Self-references (X -> X) are skipped.
    """
    mapping: Dict[int, List[int]] = {}
    current_page = None

    for line in raw_text.split("\n"):
        page_match = _PAGE_HEADER.match(line)
        if page_match:
            current_page = int(page_match.group(1))
            continue
        if current_page is None:
            continue
        for m in _CONT_ON_REF.finditer(line):
            dest = int(m.group(1))
            if dest != current_page:
                mapping.setdefault(current_page, []).append(dest)
        for m in _CONT_FROM_REF.finditer(line):
            source = int(m.group(1))
            if source != current_page:
                mapping.setdefault(source, []).append(current_page)

    # Dedupe while preserving insertion order
    return {k: list(dict.fromkeys(v)) for k, v in mapping.items()}


def _article_spans(
    page_start: int,
    page_end: int,
    cont_map: Dict[int, List[int]],
    max_hops: int = 5,
) -> List[tuple]:
    """Build list of (start, end) page-range tuples for an article, following continuation chains.

    Primary span is (page_start, page_end). Each continuation destination is added as a
    single-page span. BFS-chased up to max_hops with visited-set cycle prevention so that
    chains like {4: [18], 18: [22]} expand to [(4,5), (18,18), (22,22)] for an article
    whose index says page_start=4, page_end=5.
    """
    spans = [(page_start, page_end)]
    visited = set(range(page_start, page_end + 1))
    queue = list(range(page_start, page_end + 1))
    hops = 0
    while queue and hops < max_hops:
        next_queue = []
        for p in queue:
            for dest in cont_map.get(p, []):
                if dest in visited:
                    continue
                visited.add(dest)
                spans.append((dest, dest))
                next_queue.append(dest)
        queue = next_queue
        hops += 1
    return spans


def _parse_groq_json(raw_response: str):
    """Parse JSON from Groq response, handling markdown code fences."""
    json_str = raw_response
    fence_match = re.search(r"```(?:json)?\s*\n?(.*?)```", raw_response, re.DOTALL)
    if fence_match:
        json_str = fence_match.group(1).strip()
    return json.loads(json_str)


def pass2_segment(issue_dir: Path, meta: Dict) -> int:
    """Segment raw text into individual article .md files. Returns article count."""
    print(f"  PASS 2: Article segmentation via Groq...")

    raw_text = (issue_dir / "raw_text.txt").read_text(encoding="utf-8")

    # Build continuation map from the UNSCRUBBED text, then scrub in-memory
    # for all downstream consumers (TOC extraction, Pass 2a, Pass 2b).
    continuations_map = resolve_continuations(raw_text)
    if continuations_map:
        print(f"  Continuations detected: {continuations_map}")
    raw_text = _scrub_continuation_markers(raw_text)

    toc = _extract_toc(raw_text)

    # Step 1: Get article metadata index (small response, no body text)
    print(f"  Step 2a: Extracting article index from TOC...")
    toc_response = get_groq().chat.completions.create(
        model=GROQ_MODEL,
        max_tokens=2048,
        messages=[
            {"role": "system", "content": PASS2_TOC_SYSTEM},
            {"role": "user", "content": f"TABLE OF CONTENTS:\n{toc}\n\nFULL MAGAZINE TEXT:\n{raw_text}"},
        ],
    )

    toc_raw = (toc_response.choices[0].message.content or "").strip()
    articles_meta = _parse_groq_json(toc_raw)
    print(f"  Found {len(articles_meta)} articles in index")

    # Step 2: Extract each article body individually
    for idx, article in enumerate(articles_meta, start=1):
        title = article.get("title", "Untitled")
        author = article.get("author", "Unknown")
        page_start = article.get("page_start", 1)
        page_end = article.get("page_end", page_start)

        print(f"  Step 2b: Extracting [{idx}/{len(articles_meta)}] {title}...")

        # Get raw text for this article, stitching in any continuation spans
        spans = _article_spans(page_start, page_end, continuations_map)
        if len(spans) == 1:
            page_text = _extract_page_range(raw_text, page_start, page_end)
        else:
            parts = [_extract_page_range(raw_text, s, e) for s, e in spans]
            page_text = "\n\n[CONTINUED]\n\n".join(p for p in parts if p)
            print(f"    Stitched continuation spans: {spans}")

        body_response = get_groq().chat.completions.create(
            model=GROQ_MODEL,
            max_tokens=8192,
            messages=[
                {"role": "system", "content": PASS2_BODY_SYSTEM},
                {"role": "user", "content": (
                    f"Extract the article titled \"{title}\" by {author} "
                    f"from the following magazine text (pages {page_start}-{page_end}):\n\n"
                    f"{page_text}"
                )},
            ],
        )

        body_raw = (body_response.choices[0].message.content or "").strip()

        # Parse JSON response for body + topic_tags
        try:
            body_json = _parse_groq_json(body_raw)
            body = body_json.get("body", "")
            topic_tags = body_json.get("topic_tags", [])
        except (json.JSONDecodeError, ValueError):
            # Fallback: treat entire response as body text, no tags
            body = body_raw
            topic_tags = []

        if not body.strip():
            print(f"    ⚠ Empty body for '{title}' — skipping .md write")
            log_failure(
                meta.get("issue", ""),
                title,
                f"Empty article body (pages {page_start}-{page_end})",
                issue_dir,
            )
            continue

        # Validate tags against taxonomy
        valid_tags = [t for t in topic_tags if t in VALID_TAGS]
        invalid_tags = [t for t in topic_tags if t not in VALID_TAGS]
        if invalid_tags:
            print(f"    Removed invalid tags: {invalid_tags}")
        topic_tags = valid_tags

        tags_str = ", ".join(topic_tags) if topic_tags else ""

        # Extract Bible references via regex (fast, no LLM)
        bible_refs = _extract_bible_refs_regex(body)
        bible_refs_str = ", ".join(bible_refs) if bible_refs else ""
        if bible_refs:
            print(f"    Bible refs (regex): {len(bible_refs)} found")

        slug = slugify(title)
        filename = f"{idx:02d}_{slug}.md"

        frontmatter = (
            f"---\n"
            f"TITLE: {title}\n"
            f"AUTHOR: {author}\n"
            f"ISSUE: {meta.get('issue', '')}\n"
            f"DATE: {meta.get('date', '')}\n"
            f"PAGE_START: {page_start}\n"
            f"PAGE_END: {page_end}\n"
            f"SOURCE_TYPE: magazine_article\n"
            f"TOPIC_TAGS: {tags_str}\n"
            f"BIBLE_REFS: {bible_refs_str}\n"
            f"---\n\n"
            f"# {title}\n"
            f"*by {author}*\n\n"
            f"{body}"
        )

        (issue_dir / filename).write_text(frontmatter, encoding="utf-8")

    print(f"  {len(articles_meta)} articles segmented")
    return len(articles_meta)


# -- PASS 3: QA INSPECTION (Groq Llama 3.3 70B) -----------------------------

PASS3_SYSTEM = """You are a quality inspector for magazine article transcriptions.

Analyze this article and extract the following information:

1. first_10_words: The first 10 words of the article body (after any frontmatter/title/byline)
2. last_10_words: The last 10 words of the article body
3. word_count: Total word count of the body text
4. starts_mid_sentence: Does the body start mid-sentence or mid-word? (true/false)
5. has_duplicate_paragraphs: Are there any duplicate sentences or paragraphs? (true/false)
6. body_matches_title: Does the body text match the title and author? (true/false)
7. has_garbled_text: Are there obvious OCR errors or garbled text? (true/false)
8. issues: List of specific issues found (empty array if none)

Respond with JSON only:
{
  "first_10_words": "string",
  "last_10_words": "string",
  "word_count": number,
  "starts_mid_sentence": bool,
  "has_duplicate_paragraphs": bool,
  "body_matches_title": bool,
  "has_garbled_text": bool,
  "issues": ["list of issues found"]
}"""

# Words that suggest an article was truncated mid-sentence at the end
_CONTINUATION_WORDS = {
    "a", "an", "the", "and", "but", "or", "nor", "for", "yet", "so",
    "in", "on", "at", "to", "of", "by", "with", "from", "as", "is",
    "was", "are", "were", "be", "been", "being", "that", "this", "which",
    "who", "whom", "whose", "into", "than", "not", "its", "his", "her",
}


def _check_truncation(last_words: str) -> bool:
    """Return True if the last word suggests mid-sentence truncation."""
    words = last_words.strip().split()
    if not words:
        return False
    return words[-1].lower().rstrip(".,;:!?") in _CONTINUATION_WORDS


def pass3_qa(issue_dir: Path, issue: str = "") -> Dict[str, int]:
    """Run QA on each .md article file. Returns {pass, warn, flag} counts."""
    print(f"  PASS 3: QA inspection via Groq...")

    md_files = sorted(issue_dir.glob("*.md"))
    if not md_files:
        print(f"  No .md files found")
        return {"pass": 0, "warn": 0, "flag": 0}

    flagged_dir = issue_dir / "flagged"
    qa_results = []
    counts = {"pass": 0, "warn": 0, "flag": 0}

    for md_path in md_files:
        content = md_path.read_text(encoding="utf-8")

        try:
            response = get_groq().chat.completions.create(
                model=GROQ_MODEL,
                max_tokens=1024,
                messages=[
                    {"role": "system", "content": PASS3_SYSTEM},
                    {"role": "user", "content": content},
                ],
            )
            raw = (response.choices[0].message.content or "").strip()
            json_str = raw
            fence_match = re.search(r"```(?:json)?\s*\n?(.*?)```", raw, re.DOTALL)
            if fence_match:
                json_str = fence_match.group(1).strip()
            result = json.loads(json_str)
        except Exception as e:
            logger.warning("QA failed for %s: %s", md_path.name, e)
            result = {
                "first_10_words": "", "last_10_words": "",
                "word_count": 0, "starts_mid_sentence": False,
                "has_duplicate_paragraphs": False, "body_matches_title": True,
                "has_garbled_text": False,
                "issues": [f"QA parse error: {e}"],
            }

        # Determine status from extracted signals
        issues = list(result.get("issues", []))
        word_count = result.get("word_count", 0)
        last_words = result.get("last_10_words", "")

        # Python-side truncation check
        if _check_truncation(last_words):
            issues.append(f"Possible truncation: ends with '{last_words.split()[-1]}'")

        if result.get("starts_mid_sentence"):
            issues.append("Starts mid-sentence")
        if result.get("has_duplicate_paragraphs"):
            issues.append("Duplicate paragraphs detected")
        if not result.get("body_matches_title", True):
            issues.append("Body does not match title/author")
        if result.get("has_garbled_text"):
            issues.append("Garbled/OCR errors detected")
        if word_count and word_count < 200:
            issues.append(f"Very short article ({word_count} words)")

        # FLAG if critical issues, WARN if minor, PASS if clean
        has_truncation = result.get("starts_mid_sentence") or _check_truncation(last_words)
        has_mismatch = not result.get("body_matches_title", True)
        if has_mismatch or (word_count and word_count < 100):
            status = "FLAG"
        elif issues:
            status = "WARN"
        else:
            status = "PASS"

        result["status"] = status
        result["issues"] = issues
        result["file"] = md_path.name

        if status == "FLAG":
            flagged_dir.mkdir(exist_ok=True)
            md_path.rename(flagged_dir / md_path.name)
            counts["flag"] += 1
            # Extract title from frontmatter for logging
            title_match = re.search(r"^TITLE:\s*(.+)$", content, re.MULTILINE)
            flag_title = title_match.group(1).strip() if title_match else md_path.stem
            flag_reasons = "; ".join(issues) if issues else "Flagged by QA"
            log_failure(issue, flag_title, f"QA FLAG: {flag_reasons}", issue_dir)
        elif status == "WARN":
            issues_str = "\n".join(f"  - {i}" for i in issues)
            warning_block = f"<!-- QA WARNINGS:\n{issues_str}\n-->\n\n"
            md_path.write_text(warning_block + content, encoding="utf-8")
            counts["warn"] += 1
        else:
            counts["pass"] += 1

        qa_results.append(result)

    # Save QA report
    (issue_dir / "qa_report.json").write_text(
        json.dumps(qa_results, indent=2), encoding="utf-8"
    )

    print(f"  QA results: {counts['pass']} pass, {counts['warn']} warn, {counts['flag']} flag")
    return counts


# -- PROCESS SINGLE ISSUE ---------------------------------------------------

def process_issue(pdf_path: Path) -> str:
    """Run all 3 passes on a single PDF. Returns 'processed' or 'failed'."""
    filename = pdf_path.name
    issue_stem = pdf_path.stem
    meta = parse_issue_meta(filename)
    issue_dir = EXTRACTED_DIR / issue_stem

    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"  Issue: {meta.get('issue')}  Date: {meta.get('date')}")
    print(f"{'='*60}")

    init_tracker()

    try:
        # Pass 1: Vision extraction
        page_count = pass1_extract(pdf_path, issue_dir)

        if page_count < 0:
            # Missing pages after all retries — abort before Pass 2
            update_tracker_row(filename, {
                "Issue": meta.get("issue"),
                "Year": meta.get("year"),
                "Pass1": "aborted_missing_pages",
                "Status": "aborted: missing pages",
            })
            PDF_FAILED_DIR.mkdir(parents=True, exist_ok=True)
            failed_dest = PDF_FAILED_DIR / filename
            pdf_path.rename(failed_dest)
            print(f"  PDF moved to failed queue: {failed_dest}")
            return "failed"

        update_tracker_row(filename, {
            "Issue": meta.get("issue"),
            "Year": meta.get("year"),
            "Pages": page_count,
            "Pass1": "complete",
            "Status": "pass1_done",
        })

        # Pass 2: Article segmentation
        article_count = pass2_segment(issue_dir, meta)
        update_tracker_row(filename, {
            "Articles": article_count,
            "Pass2": "complete",
            "Status": "pass2_done",
        })

        # Pass 3: QA inspection
        qa_counts = pass3_qa(issue_dir, issue=meta.get("issue", ""))
        update_tracker_row(filename, {
            "Pass3": "complete",
            "Pass_Count": qa_counts["pass"],
            "Warn_Count": qa_counts["warn"],
            "Flag_Count": qa_counts["flag"],
            "Status": "complete",
        })

        # Move PDF into the extracted issue folder alongside the .md files
        # so it rides with the folder through 03_approved/ until ingest archives it
        dest = issue_dir / filename
        pdf_path.rename(dest)
        print(f"  PDF moved to: {dest}")

        print(f"  DONE: {filename}")
        return "processed"

    except Exception as e:
        logger.exception("Failed processing %s", filename)
        update_tracker_row(filename, {"Status": f"failed: {str(e)[:100]}"})
        try:
            PDF_FAILED_DIR.mkdir(parents=True, exist_ok=True)
            failed_dest = PDF_FAILED_DIR / filename
            pdf_path.rename(failed_dest)
            print(f"  PDF moved to failed queue: {failed_dest}")
        except Exception:
            logger.exception("Could not move failed PDF %s to %s", filename, PDF_FAILED_DIR)
        return "failed"


# -- MAIN --------------------------------------------------------------------

def run(time_limit_min=None, max_issues=None):
    """Scan 01_to_extract/ and process all PDFs.
    If time_limit_min is set, stop after the current PDF once the limit is reached.
    If max_issues is set, stop after that many issues have been attempted (processed + failed)."""
    TO_EXTRACT_DIR.mkdir(parents=True, exist_ok=True)
    EXTRACTED_DIR.mkdir(parents=True, exist_ok=True)

    pdfs = sorted(TO_EXTRACT_DIR.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {TO_EXTRACT_DIR}")
        return

    print(f"Found {len(pdfs)} PDF(s) to process")
    if time_limit_min:
        print(f"Time limit: {time_limit_min} minutes")
    if max_issues:
        print(f"Max issues: {max_issues}")

    start_time = time.time()
    processed = failed = 0
    for pdf_path in pdfs:
        if time_limit_min:
            elapsed_min = (time.time() - start_time) / 60
            if elapsed_min >= time_limit_min:
                print(f"\nTime limit reached ({elapsed_min:.1f} min) — stopping.")
                break
        if max_issues is not None and (processed + failed) >= max_issues:
            print(f"\nMax issues reached ({max_issues}) — stopping.")
            break

        result = process_issue(pdf_path)
        if result == "processed":
            processed += 1
        else:
            failed += 1

    elapsed = (time.time() - start_time) / 60
    print(f"\n{'='*60}")
    print(f"Done. {processed} processed, {failed} failed. ({elapsed:.1f} min)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="New Wine Magazine extraction pipeline")
    parser.add_argument("--time-limit", type=float, default=None,
                        help="Stop after this many minutes (finishes current PDF first)")
    parser.add_argument("--max-issues", type=int, default=None,
                        help="Stop after this many issues have been attempted (processed + failed)")
    args = parser.parse_args()
    run(time_limit_min=args.time_limit, max_issues=args.max_issues)
