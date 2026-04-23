#!/usr/bin/env python3
"""
scrape_youtube.py — Rhemata YouTube Transcript Scraper

Pipeline:
  1. Load env vars from .env
  2. Read active channels from tracker
  3. Fetch video list via yt-dlp
  4. Download audio via yt-dlp
  5. Transcribe with local Whisper (base model)
  6. Write structured .txt to sources/youtube/raw/
  7. Update tracker immediately after each video
"""

import os
import re
import sys
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import whisper
from supabase import create_client

# ── Paths ─────────────────────────────────────────────────────────────────────
ENV_PATH         = Path("/Users/alexwhitley/Desktop/rhemata/backend/app/.env")
TRACKER_PATH     = Path("/Users/alexwhitley/Desktop/rhemata/sources/youtube/youtube_tracker.xlsx")
OUTPUT_DIR       = Path("/Users/alexwhitley/Desktop/rhemata/sources/youtube/raw")
COOKIES_PATH     = Path("/Users/alexwhitley/Desktop/rhemata/scripts/youtube_cookies.txt")
WHISPER_MODEL    = "base"

SKIP_TITLE_KEYWORDS = ["#shorts", "trailer", "promo", "highlights"]

# ── Test / throttle controls ──────────────────────────────────────────────────
TEST_MODE            = False  # False = run all active channels
MAX_TRANSCRIPTS      = 10     # Stop after this many transcripts saved


# ── Environment ───────────────────────────────────────────────────────────────

def load_env(path: Path):
    """Load key=value pairs from a .env file into os.environ."""
    if not path.exists():
        print(f"⚠  .env not found at {path} — ANTHROPIC_API_KEY must already be set")
        return
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


# ── yt-dlp helpers ────────────────────────────────────────────────────────────

def find_ytdlp():
    candidates = [
        shutil.which("yt-dlp"),
        os.path.expanduser("~/Library/Python/3.9/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.10/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.11/bin/yt-dlp"),
        os.path.expanduser("~/Library/Python/3.12/bin/yt-dlp"),
        os.path.expanduser("~/.local/bin/yt-dlp"),
    ]
    for c in candidates:
        if c and os.path.isfile(c):
            return c
    return None


def _ytdlp_base_args(ytdlp: str) -> list:
    """Common yt-dlp flags: force IPv4, set player client, cookies."""
    args = [
        ytdlp, "-4",
        "--extractor-args", "youtube:player_client=android_vr,web_safari",
    ]
    if COOKIES_PATH.exists():
        args += ["--cookies", str(COOKIES_PATH)]
    return args


def get_video_list(ytdlp: str, channel_url: str) -> list:
    cmd = _ytdlp_base_args(ytdlp) + [
        "--flat-playlist",
        "--print", "%(id)s|%(title)s|%(upload_date)s|%(duration)s",
    ]
    cmd.append(channel_url)
    result = subprocess.run(cmd, capture_output=True, text=True)
    videos = []
    for line in result.stdout.strip().splitlines():
        if "|" not in line:
            continue
        parts = line.split("|", 3)
        if len(parts) != 4:
            continue
        vid_id, title, date, duration = parts
        vid_id = vid_id.strip()
        if not vid_id:
            continue
        duration_sec = int(duration.strip()) if duration.strip().isdigit() else 0
        videos.append({
            "id":           vid_id,
            "url":          f"https://www.youtube.com/watch?v={vid_id}",
            "title":        title.strip(),
            "date":         date.strip(),
            "duration_sec": duration_sec,
        })
    return videos


AUDIO_EXTENSIONS = {".m4a", ".mp3", ".opus", ".ogg", ".webm", ".wav"}


def download_audio(ytdlp: str, video_url: str, tmp_dir: str) -> str:
    """Download audio from a YouTube video via yt-dlp. Returns path to audio file or None."""
    out_template = os.path.join(tmp_dir, "%(id)s.%(ext)s")
    cmd = _ytdlp_base_args(ytdlp) + [
        "-x",
        "-o", out_template,
        video_url,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        error_lines = [l for l in result.stderr.splitlines() if l.strip().startswith("ERROR")]
        error_msg = error_lines[-1] if error_lines else result.stderr.strip().splitlines()[-1]
        print(f"     yt-dlp audio error: {error_msg[:200]}")
        return None
    # Find the downloaded audio file (any format — Whisper handles all of them)
    for f in os.listdir(tmp_dir):
        if os.path.splitext(f)[1].lower() in AUDIO_EXTENSIONS:
            return os.path.join(tmp_dir, f)
    return None


def transcribe_audio(model, audio_path: str):
    """Transcribe an audio file with Whisper. Returns text or None."""
    try:
        result = model.transcribe(audio_path, language="en")
        return result.get("text", "").strip() or None
    except Exception as e:
        print(f"     Whisper error: {type(e).__name__}: {str(e)[:120]}")
        return None


# ── File helpers ──────────────────────────────────────────────────────────────

def format_date(date_str: str) -> str:
    if date_str and len(date_str) == 8 and date_str.isdigit():
        return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    return date_str or "unknown"


def make_filename(date_str: str, handle: str, title: str) -> str:
    safe_title = re.sub(r'[^\w\s]', '', title.lower())
    safe_title = re.sub(r'\s+', '_', safe_title.strip())[:60]
    handle_clean = handle.lstrip('@')
    date_clean = date_str if (date_str and date_str != "unknown") else "00000000"
    return f"{date_clean}_{handle_clean}_{safe_title}.txt"


def write_transcript_file(path: Path, channel: dict, video: dict,
                          cleaned_text: str):
    pub_date = format_date(video["date"])
    duration_min = round(video["duration_sec"] / 60, 1)
    header = (
        f"TITLE: {video['title']}\n"
        f"SPEAKER: {channel['speaker']}\n"
        f"CHANNEL: {channel['name']}\n"
        f"URL: {video['url']}\n"
        f"SOURCE_URL: {video['url']}\n"
        f"PUBLISHED: {pub_date}\n"
        f"DURATION_MIN: {duration_min}\n"
        f"SOURCE_TYPE: sermon\n"
        f"---\n\n"
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(header + cleaned_text, encoding="utf-8")


# ── Supabase dedupe ──────────────────────────────────────────────────────────

def _strip_punctuation(s):
    """Strip all punctuation for fuzzy title matching."""
    return re.sub(r'[^\w\s]', '', s).strip()


def init_supabase():
    """Create Supabase client from env vars. Returns client or None."""
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if url and key:
        return create_client(url, key)
    print("  ⚠  SUPABASE_URL or SUPABASE_SERVICE_KEY not set — skipping dedupe")
    return None


def already_in_supabase(db, title, speaker):
    """Check if a document with a matching title and author exists in Supabase."""
    if not db:
        return False
    clean_title = _strip_punctuation(title)
    try:
        result = (
            db.table("documents")
            .select("id, title")
            .ilike("author", f"%{speaker}%")
            .execute()
        )
        for doc in result.data:
            if _strip_punctuation(doc.get("title", "")) == clean_title:
                return True
    except Exception as e:
        print(f"     ⚠  Supabase check failed: {e}")
    return False


# ── Tracker helpers ───────────────────────────────────────────────────────────

def load_tracker():
    return openpyxl.load_workbook(TRACKER_PATH)


def get_scraped_urls(wb) -> set:
    ws = wb["Videos"]
    scraped = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        url, scraped_flag = row[4], row[7]
        if url and str(scraped_flag).strip() == "Yes":
            scraped.add(str(url).strip())
    return scraped


def get_active_channels(wb) -> list:
    ws = wb["Channels"]
    channels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[4] and str(row[4]).strip().lower() == "yes":
            channels.append({
                "handle":  str(row[0]).strip(),
                "name":    str(row[1]).strip(),
                "url":     str(row[2]).strip(),
                "speaker": str(row[3]).strip() if row[3] else str(row[1]).strip(),
            })
    return channels


def log_row(wb, channel, video, scraped, filepath, word_count, note=""):
    pub_date = format_date(video["date"])
    duration_min = round(video["duration_sec"] / 60, 1)
    wb["Videos"].append([
        channel["handle"],
        channel["name"],
        channel["speaker"],
        video["title"],
        video["url"],
        pub_date,
        duration_min,
        scraped,
        filepath,
        word_count or "",
        datetime.now().strftime("%Y-%m-%d"),
        note,
    ])


def save_tracker(wb):
    wb.save(TRACKER_PATH)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    load_env(ENV_PATH)

    ytdlp = find_ytdlp()
    if not ytdlp:
        print("✗ yt-dlp not found. Run: pip3 install yt-dlp")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Loading Whisper model '{WHISPER_MODEL}'...")
    model = whisper.load_model(WHISPER_MODEL)
    print(f"Whisper:  {WHISPER_MODEL} model loaded")

    db = init_supabase()
    print(f"yt-dlp:   {ytdlp}")
    print(f"Cookies:  {COOKIES_PATH}{' ✓' if COOKIES_PATH.exists() else ' ✗ (not found — may be IP-blocked)'}")
    print(f"Output:   {OUTPUT_DIR}")
    print(f"Tracker:  {TRACKER_PATH}\n")

    wb = load_tracker()
    channels = get_active_channels(wb)
    scraped_urls = get_scraped_urls(wb)

    print(f"{len(channels)} active channel(s) | {len(scraped_urls)} already scraped | max transcripts: {MAX_TRANSCRIPTS}\n")

    stats = {"new": 0, "skipped": 0, "errors": 0, "supabase_skip": 0}
    done = False

    for ch in channels:
        if done:
            break
        print(f"{'═' * 64}")
        print(f"  {ch['name']} ({ch['handle']})")

        try:
            videos = get_video_list(ytdlp, ch["url"])
        except Exception as e:
            print(f"  ✗ Failed to fetch video list: {e}")
            stats["errors"] += 1
            continue

        print(f"  {len(videos)} video(s) in playlist\n")

        for vid in videos:
            if stats["new"] >= MAX_TRANSCRIPTS:
                print(f"\n  Reached {MAX_TRANSCRIPTS} transcripts — stopping.")
                done = True
                break

            url   = vid["url"]
            title = vid["title"]
            dur   = vid["duration_sec"]

            # Skip already scraped
            if url in scraped_urls:
                stats["skipped"] += 1
                continue

            # Skip short videos
            if dur > 0 and dur < 60:
                print(f"  ⏭  {title[:60]} ({dur}s — too short)")
                log_row(wb, ch, vid, "No", "", 0, "Skipped: too short")
                scraped_urls.add(url)
                save_tracker(wb)
                stats["skipped"] += 1
                continue

            # Skip by title keyword
            if any(kw in title.lower() for kw in SKIP_TITLE_KEYWORDS):
                print(f"  ⏭  {title[:60]} — title filtered")
                stats["skipped"] += 1
                continue

            print(f"  → {title[:70]} [{stats['new']}/{MAX_TRANSCRIPTS} saved]")

            # Supabase dedupe — skip if already ingested
            if already_in_supabase(db, title, ch["speaker"]):
                print(f"     SKIP — already in Supabase: {title[:70]}")
                stats["supabase_skip"] += 1
                continue

            # Download audio → transcribe with Whisper
            with tempfile.TemporaryDirectory() as tmp_dir:
                print(f"     Downloading audio...")
                audio_path = download_audio(ytdlp, url, tmp_dir)
                if not audio_path:
                    print(f"     ✗ Audio download failed")
                    log_row(wb, ch, vid, "No", "", 0, "Audio download failed")
                    scraped_urls.add(url)
                    save_tracker(wb)
                    stats["errors"] += 1
                    continue

                print(f"     Transcribing with Whisper...")
                raw_text = transcribe_audio(model, audio_path)

            if not raw_text:
                print(f"     ✗ Whisper transcription failed")
                log_row(wb, ch, vid, "No", "", 0, "Whisper transcription failed")
                scraped_urls.add(url)
                save_tracker(wb)
                stats["errors"] += 1
                continue

            # Write transcript file
            fname    = make_filename(vid["date"], ch["handle"], title)
            out_path = OUTPUT_DIR / fname
            try:
                write_transcript_file(out_path, ch, vid, raw_text)
                word_count = len(raw_text.split())
                print(f"     ✓  {fname} ({word_count:,} words)")
                log_row(wb, ch, vid, "Yes", str(out_path), word_count, "")
                scraped_urls.add(url)
                stats["new"] += 1
            except Exception as e:
                print(f"     ✗ Write error: {e}")
                log_row(wb, ch, vid, "No", "", 0, f"Write error: {e}")
                scraped_urls.add(url)
                stats["errors"] += 1

            save_tracker(wb)

    print(f"\n{'═' * 64}")
    print(f"Done.")
    print(f"  ✓ New transcripts   : {stats['new']}")
    print(f"  ⏭  Skipped          : {stats['skipped']}")
    print(f"  ⏭  Already in DB    : {stats['supabase_skip']}")
    print(f"  ✗ Errors            : {stats['errors']}")
    print(f"  Output: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
